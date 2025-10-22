"""
Serviço de Exportação de Dados
Suporta Excel, CSV e PDF para relatórios administrativos
"""
import csv
import io
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas


class ExportService:
    """
    Serviço centralizado para exportação de dados em múltiplos formatos
    """
    
    @staticmethod
    def _sanitize_value(value: Any) -> str:
        """Converte valor para string segura para exportação"""
        if value is None:
            return ''
        if isinstance(value, (int, float, Decimal)):
            return str(value)
        if isinstance(value, datetime):
            return value.strftime('%d/%m/%Y %H:%M')
        if isinstance(value, bool):
            return 'Sim' if value else 'Não'
        return str(value)
    
    @staticmethod
    def _calculate_optimal_col_widths(
        headers: Dict[str, str],
        data: List[Dict[str, Any]],
        available_width: float
    ) -> List[float]:
        """
        Calcula larguras ótimas de colunas baseado no conteúdo
        Retorna lista de larguras em pontos
        """
        col_keys = list(headers.keys())
        col_widths = []
        
        for key in col_keys:
            # Padrões específicos por tipo de campo
            key_lower = key.lower()
            
            if 'address' in key_lower or 'endereco' in key_lower or 'endereço' in key_lower:
                weight = 0.25  # 25% para endereços (campo mais largo)
            elif 'email' in key_lower:
                weight = 0.15  # 15% para emails
            elif 'description' in key_lower or 'descricao' in key_lower or 'descrição' in key_lower:
                weight = 0.20  # 20% para descrições
            elif 'number' in key_lower or 'numero' in key_lower or 'número' in key_lower:
                weight = 0.10  # 10% para números de pedido/referência
            elif 'status' in key_lower:
                weight = 0.08  # 8% para status
            elif 'date' in key_lower or 'data' in key_lower:
                weight = 0.12  # 12% para datas
            elif 'total' in key_lower or 'valor' in key_lower or 'price' in key_lower or 'preco' in key_lower or 'preço' in key_lower:
                weight = 0.10  # 10% para valores monetários
            elif 'name' in key_lower or 'nome' in key_lower:
                weight = 0.15  # 15% para nomes
            elif 'tracking' in key_lower or 'rastreamento' in key_lower:
                weight = 0.12  # 12% para códigos de rastreamento
            elif 'method' in key_lower or 'metodo' in key_lower or 'método' in key_lower:
                weight = 0.10  # 10% para métodos de pagamento/envio
            else:
                weight = 0.12  # 12% padrão
            
            col_widths.append(weight)
        
        # Normalizar para somar 100% do espaço disponível
        total_weight = sum(col_widths)
        col_widths = [(w / total_weight) * available_width for w in col_widths]
        
        return col_widths
    
    # ========================================
    # EXPORTAÇÃO PARA EXCEL
    # ========================================
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        filename: str,
        title: str = None
    ) -> HttpResponse:
        """
        Exporta dados para Excel (.xlsx) com formatação profissional
        
        Args:
            data: Lista de dicionários com os dados
            headers: Dicionário {chave: nome_coluna}
            filename: Nome do arquivo sem extensão
            title: Título opcional da planilha
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=14, color="1F4788")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        
        # Adicionar título se fornecido
        if title:
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = title_font
            title_cell.alignment = title_alignment
            row_num = 2
            
            # Data de geração
            ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
            date_cell = ws['A2']
            date_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            date_cell.alignment = Alignment(horizontal="center")
            row_num = 4
        
        # Cabeçalhos
        for col_num, (key, header) in enumerate(headers.items(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Dados
        for row_data in data:
            row_num += 1
            for col_num, key in enumerate(headers.keys(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = ExportService._sanitize_value(row_data.get(key, ''))
                cell.border = border_style
                
                # Alinhamento baseado no tipo
                value = row_data.get(key)
                if isinstance(value, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Auto-ajustar largura das colunas
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Gerar resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response
    
    # ========================================
    # EXPORTAÇÃO PARA CSV
    # ========================================
    
    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        filename: str
    ) -> HttpResponse:
        """
        Exporta dados para CSV
        
        Args:
            data: Lista de dicionários com os dados
            headers: Dicionário {chave: nome_coluna}
            filename: Nome do arquivo sem extensão
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        # BOM para Excel reconhecer UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        
        # Cabeçalhos
        writer.writerow(headers.values())
        
        # Dados
        for row_data in data:
            row = [ExportService._sanitize_value(row_data.get(key, '')) for key in headers.keys()]
            writer.writerow(row)
        
        return response
    
    # ========================================
    # EXPORTAÇÃO PARA PDF
    # ========================================
    
    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        filename: str,
        title: str = "Relatório",
        orientation: str = 'landscape'
    ) -> HttpResponse:
        """
        Exporta dados para PDF com tabela formatada
        
        Args:
            data: Lista de dicionários com os dados
            headers: Dicionário {chave: nome_coluna}
            filename: Nome do arquivo sem extensão
            title: Título do relatório
            orientation: 'portrait' ou 'landscape'
        """
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        # Buffer
        buffer = io.BytesIO()
        
        # Configurar página
        pagesize = landscape(A4) if orientation == 'landscape' else A4
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Container para elementos
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Título
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))
        
        # Preparar dados da tabela com Paragraphs para quebra de linha automática
        normal_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='LTR',
            alignment=0  # Left
        )
        
        # Cabeçalhos como Paragraphs
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.whitesmoke,
            alignment=1,  # Center
            fontName='Helvetica-Bold'
        )
        
        table_data = [[Paragraph(str(header), header_style) for header in headers.values()]]
        
        # Dados com Paragraphs para quebra automática
        for row_data in data:
            row = []
            for key in headers.keys():
                cell_value = ExportService._sanitize_value(row_data.get(key, ''))
                # Usar Paragraph para permitir quebra de linha
                cell_paragraph = Paragraph(str(cell_value), normal_style)
                row.append(cell_paragraph)
            table_data.append(row)
        
        # Calcular larguras de coluna dinamicamente usando método otimizado
        available_width = pagesize[0] - 60  # Margem total (30 left + 30 right)
        col_widths = ExportService._calculate_optimal_col_widths(headers, data, available_width)
        
        # Criar tabela com larguras personalizadas
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Estilo da tabela com VALIGN para alinhamento vertical
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Dados
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),  # Alinhamento vertical no topo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1F4788')),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            
            # Quebra de linha automática
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        elements.append(table)
        
        # Footer com número de registros
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Total de registros: {len(data)}",
            styles['Normal']
        ))
        
        # Gerar PDF
        doc.build(elements)
        
        # Retornar resposta
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    
    # ========================================
    # MÉTODOS DE CONVENIÊNCIA
    # ========================================
    
    @staticmethod
    def export_data(
        data: List[Dict[str, Any]],
        headers: Dict[str, str],
        format_type: str,
        filename: str,
        title: str = None
    ) -> HttpResponse:
        """
        Método unificado para exportar em qualquer formato
        
        Args:
            data: Dados a exportar
            headers: Cabeçalhos das colunas
            format_type: 'excel', 'csv', ou 'pdf'
            filename: Nome base do arquivo
            title: Título para Excel e PDF
        """
        if format_type == 'excel':
            return ExportService.export_to_excel(data, headers, filename, title)
        elif format_type == 'csv':
            return ExportService.export_to_csv(data, headers, filename)
        elif format_type == 'pdf':
            return ExportService.export_to_pdf(data, headers, filename, title or filename)
        else:
            raise ValueError(f"Formato não suportado: {format_type}")


def get_export_service() -> ExportService:
    """Retorna instância do serviço de exportação"""
    return ExportService()
